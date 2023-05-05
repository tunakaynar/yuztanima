import cv2
from simple_facerec import SimpleFacerec
from datetime import datetime
from datetime import date
import pandas as pd
import openpyxl
import pypyodbc
import os
import sys
import os
from playsound import playsound
import time

arguman = "Bizimİnternet"

# try:
#      arguman = sys.argv[1]
# except:
#      print('Arguman ver!!!')
#      exit()

today = date.today()
day = today.strftime("%b-%d-%Y")
yoklama_csv = arguman + "-yoklama-" + day + ".csv"
yoklama_xlsx = arguman + "-yoklama-" + day + ".xlsx"

def yoklamayi_temizle():
    if os.path.exists(yoklama_csv):
        os.remove(yoklama_csv)

    if os.path.exists(yoklama_xlsx):
        os.remove(yoklama_xlsx)

    db = pypyodbc.connect(
        'Driver={SQL Server};'
        'Server=DESKTOP-LE5T4BA\SQLEXPRESS;'
        'Database=DbOkul;'
        'UID=;'
        'PWD=;'
    )
    imlec = db.cursor()
    query = f"DELETE FROM YOKLAMA1"
    sonuc = imlec.execute(query)
    db.commit()

def dosyalari_ayarla():
    yoklamayi_temizle()
    dosya = open(yoklama_csv, "a")
    dosya.write("Ad, Saat")
    dosya.close()


def yoklama_sql_tablosuna_kaydet(numara):
    db = pypyodbc.connect(
        'Driver={SQL Server};'
        'Server=DESKTOP-LE5T4BA\SQLEXPRESS;'
        'Database=DbOkul;'
        'UID=;'
        'PWD=;'
    )
    imlec = db.cursor()
    query = f"DECLARE @table_count INT WITH ogrenciler as ( SELECT  O.OGRID, O.DERSID, D.DERSADI FROM OGRDERS O INNER JOIN DERSLER D ON  O.DERSID = D.ID WHERE O.DERSID = 1 AND D.DERSADI = 'FİZİK' ) select @table_count = COUNT(*) from ogrenciler  IF @table_count > 0 INSERT INTO YOKLAMA1 (NUMARA, AD, BOLUM, FAKULTE, TARIH, DERS) SELECT  OG.NUMARA, OG.AD, OG.BOLUM, OG.FAKULTE, GETDATE(), '{arguman}' FROM OGRENCI OG WHERE OG.NUMARA = '{numara}'"
    sonuc = imlec.execute(query)
    db.commit()


def yoklamayaYaz(numara):
    with open(yoklama_csv, 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])

        if numara not in nameList:
            yoklama_sql_tablosuna_kaydet(numara)
            now = datetime.now()
            dtString = now.strftime('%H:%M:%S')
            f.writelines(f'\n{numara},{dtString}')

            sound_file = 'yoklama_alindi.mp3'
            if os.path.exists(sound_file):
                playsound(sound_file)
            else:
                print(f"{sound_file} adlı ses dosyası bulunamadı.")

dosyalari_ayarla()

# Klasör tarama işleminin yapılması
sfr = SimpleFacerec()
sfr.load_encoding_images("images/")

# Kamera Açılması
# rtsp://<admin>:<pass>@ip:port/cam/realmonitor?channel=1&subtype=1
ip='rtsp://admin:biziminternet1@10.29.214.120'


cap = cv2.VideoCapture(1)






while True:


    ret, frame = cap.read()


    # Yüz Belirleme
    face_locations, face_names = sfr.detect_known_faces(frame)
    for face_loc, name in zip(face_locations, face_names):
        y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]

        cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 4)

        yoklamayaYaz(name)

    cv2.imshow("Camera", frame)

    key = cv2.waitKey(1)
    if key == 27:
        break

    data = pd.read_csv(arguman + "-yoklama-" + day + ".csv")

    wb = openpyxl.Workbook()
    sayfa = wb.active

    a2 = len(data)  ### toplam satır sayısı
    a3 = len(data.columns)  ### toplam sütun sayısı
    print('satır uzunluğu: ', a2)
    print('sütun sayısı: ', a3)

    for x in range(a3):  ### sütun başlıklarını yazdırma döngüsü
        c = x + 1
        sayfa.cell(row=1, column=c).value = data.columns[x]
        

    for x in range(a2):  ### tüm satırlardaki verileri excele yazdırma döngüsü
        for y in range(a3):
            r = x + 2
            c = y + 1
            sayfa.cell(row=r, column=c).value = data.iat[x, y]

    wb.save(arguman + "-yoklama-" + day + ".xlsx")

    print('İşlem başarıyla tamamlandı. Excel dosyanız oluşturuldu')
    

cap.release()
cv2.destroyAllWindows()